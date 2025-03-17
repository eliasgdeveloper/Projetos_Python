import os
from openpyxl import load_workbook

# Caminho do arquivo Excel
caminhoArquivo = r"C:\Users\elias\Desktop\Pasta2.xlsx"

# Verifica se o arquivo existe
if os.path.exists(caminhoArquivo):
    # Carregar a planilha Excel
    planilha = load_workbook(caminhoArquivo)

    # Obter os nomes das abas
    nome_aba = planilha.sheetnames
    print(f"Abas disponíveis: {nome_aba}")

    # Verificar se existem pelo menos duas abas
    if len(nome_aba) >= 2:
        # Selecionar a primeira aba
        aba1 = planilha[nome_aba[0]]

        # Inserir valores na aba 1
        aba1["A1"] = "NOME"
        aba1["B1"] = "DATA"
        print(f"Valores 'NOME' e 'DATA' adicionados na aba: {nome_aba[0]}")

        # Selecionar a segunda aba
        aba2 = planilha[nome_aba[1]]

        # Inserir valores na aba 2
        aba2["A1"] = "NOME"
        aba2["B1"] = "DATA"
        print(f"Valores 'NOME' e 'DATA' adicionados na aba: {nome_aba[1]}")

        # Salvar as alterações
        planilha.save(caminhoArquivo)
        print(f"Arquivo salvo com sucesso!")
    else:
        print("Erro: O arquivo precisa ter pelo menos duas abas.")
else:
    print(f"O arquivo '{caminhoArquivo}' não foi encontrado.")




    num1 = int(input("Digite o primeiro número: "))
    num2 = int(input("Digite o segundo número: "))
    soma = num1 + num2
    print("A soma é:", soma)

